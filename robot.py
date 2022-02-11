from asyncio import sleep
from http import server
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.descriptors.base import Length
# import pyautogui
# import win32clipboard
from openpyxl import Workbook
from openpyxl import load_workbook,worksheet
from random import randint
wb=Workbook()
wb = load_workbook(filename = 'orginal.xlsx')
wc = load_workbook(filename = 'duporginal.xlsx')
sheet_ranges = wb['Sheet1']
ws=wc.active
import  os
import time


path=os.path.dirname(os.path.abspath(__file__))
address= os.path.join(path ,'chromedriver.exe' )

driver=webdriver.Chrome(executable_path= address)
driver.get('http://tvqc-tportal.moe.gov.ir')
time.sleep(2) 

def rand_value(l):
    if((l%10)==0):
        randvalue=randint(3,5)
    else:
        randvalue=randint(2,5)
    return randvalue



def login(user):
    fielduser = driver.find_element_by_name('txt_username')
    fieldpass = driver.find_element_by_name('txt_password')
    fielduser.send_keys('p_'+user)
    fieldpass.send_keys(user)
    enterbutton=driver.find_element_by_xpath('//*[@id="btn_login"]').click()   

def check_url(url):
  current_url = driver.current_url
  result=(lambda url:True if(url==current_url) else False)(url)
  return result

def link_to_courses():
     time.sleep(2)
     linkcourses=driver.find_element_by_xpath('//*[@id="main-menu-wrapper"]/ul/li[1]').click()

def search_course(waite,course_id):    
    time.sleep(2)
    search_field=driver.find_element_by_xpath('//*[@id="div_box_data_1"]/div/div[2]/div/div[2]/div[1]/div[3]/input')
    time.sleep(2)
    search_field.clear()
    search_field.send_keys(course_id)
    time.sleep(2)
    button=driver.find_elements(By.XPATH, '//*[@id="tbl_1"]/tbody/tr/td[16]/a')
    if  button:
        button[0].click()
        time.sleep(2)
        back_button=driver.find_elements(By.XPATH, '//*[@id="btn_100_back"]')
        if back_button:
            back_button[0].click() 
            result=False
        else:
            result=True
    else:
        result=False
    return result
    
def submit_vote():
    driver.find_element_by_xpath('//*[@id="ddl_question_91923"]').send_keys(rand_value(l))
    driver.find_element_by_xpath('//*[@id="ddl_question_91924"]').send_keys(rand_value(l))
    driver.find_element_by_xpath('//*[@id="ddl_question_91925"]').send_keys(rand_value(l))
    driver.find_element_by_xpath('//*[@id="ddl_question_91926"]').send_keys(rand_value(l))
    driver.find_element_by_xpath('//*[@id="ddl_question_91927"]').send_keys(rand_value(l))
    driver.find_element_by_xpath('//*[@id="ddl_question_91928"]').send_keys(rand_value(l))
    driver.find_element_by_xpath('//*[@id="ddl_question_91929"]').send_keys(rand_value(l))
    driver.find_element_by_xpath('//*[@id="ddl_question_91930"]').send_keys(rand_value(l))
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="btn_save"]').click() 
    time.sleep(2)
    driver.find_element_by_xpath('/html/body/div[5]/div/div[3]/button[1]').click()
    time.sleep(2)

def logout():
    logout=driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[3]/ul[1]/li/a').click()

     
# main.............................
Lengthdata=len(sheet_ranges['L'])
l=1
while l<=Lengthdata:
    user1=sheet_ranges['M'+str(l)].value
    user2=sheet_ranges['M'+str(l+1)].value
    course1=sheet_ranges['L'+str(l)].value
    course2=sheet_ranges['L'+str(l+1)].value 
    print(user1)   
    print(course1)   
    print(l)
    print('........................')   
    login(str(user1)) 
    time.sleep(5)
    result=driver.find_elements(By.XPATH,'//*[@id="main-menu-wrapper"]/ul/li[1]') 
    # login_condition
    if  result:
        link_to_courses()
        search_course_res=search_course(10,course1)
        if (search_course_res==True):
            submit_vote()
        else:
            j=1
        while user1==user2:
            search_course_res2=search_course(10,course2)
            if (search_course_res2==True):
                submit_vote()
                user1=sheet_ranges['M'+str(l+j)].value
                user2=sheet_ranges['M'+str(l+j+1)].value
                course1=sheet_ranges['L'+str(l+j)].value
                course2=sheet_ranges['L'+str(l+j+1)].value
                j=j+1
            else:
                user1=sheet_ranges['M'+str(l+j)].value
                user2=sheet_ranges['M'+str(l+j+1)].value
                course1=sheet_ranges['L'+str(l+j)].value
                course2=sheet_ranges['L'+str(l+j+1)].value
                j=j+1
        logout()
        l=l+j
    else:
       driver.close()
       driver=webdriver.Chrome(executable_path= address)
       driver.get('http://tvqc-tportal.moe.gov.ir')
       time.sleep(2)
       l=l+1
driver.close()